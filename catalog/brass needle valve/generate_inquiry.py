import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import datetime
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Inquiry"

# ── helpers ──────────────────────────────────────────────────────────────────
navy   = "1F3864"
gold   = "C9A84C"
white  = "FFFFFF"
light  = "EAF0FB"
mid    = "D6E0F5"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value="", bold=False, size=11, color="000000",
         bg=None, align="left", wrap=False, italic=False, border_on=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=color, italic=italic,
                  name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    if bg:
        c.fill = fill(bg)
    if border_on:
        c.border = border()
    return c

# ── column widths ─────────────────────────────────────────────────────────────
col_widths = {1: 3, 2: 22, 3: 55, 4: 20, 5: 3}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── row heights ───────────────────────────────────────────────────────────────
for r in range(1, 50):
    ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height  = 8   # top margin
ws.row_dimensions[2].height  = 48  # banner
ws.row_dimensions[3].height  = 8   # gap
ws.row_dimensions[18].height = 8   # gap before table
ws.row_dimensions[26].height = 55  # question row

# ── BANNER ────────────────────────────────────────────────────────────────────
ws.merge_cells("B2:D2")
c = ws["B2"]
c.value       = "PRODUCT INQUIRY  ·  UL CERTIFICATION REQUEST"
c.font        = Font(bold=True, size=18, color=white, name="Calibri")
c.fill        = fill(navy)
c.alignment   = Alignment(horizontal="center", vertical="center")

# thin gold line under banner
ws.merge_cells("B3:D3")
ws["B3"].fill = fill(gold)
ws.row_dimensions[3].height = 4

# ── HEADER META ───────────────────────────────────────────────────────────────
meta = [
    ("Date",        datetime.date.today().strftime("%B %d, %Y")),
    ("From",        "ManoEng / Procurement Team"),
    ("Subject",     "UL Certification Availability Inquiry"),
]
row = 4
for label, value in meta:
    ws.row_dimensions[row].height = 20
    cell(ws, row, 2, label,  bold=True, size=10, color=white, bg=navy,
         align="right", border_on=True)
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row,   end_column=4)
    cell(ws, row, 3, value, size=10, bg=light, border_on=True)
    row += 1

# gap
row += 1  # row 8

# ── SECTION: PRODUCT DETAILS ─────────────────────────────────────────────────
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
c = ws.cell(row=row, column=2,
            value="  PRODUCT REFERENCE")
c.font      = Font(bold=True, size=11, color=white, name="Calibri")
c.fill      = fill(gold)
c.alignment = Alignment(horizontal="left", vertical="center")
row += 1

specs = [
    ("Listing Title",
     'High Quality C-Way 1/4\'\' Brass Mini Needle Valve 2-Way 400 psi '
     'High Temp Wheel Handle for Water Media  OEM Acceptable'),
    ("Valve Type",       "Needle Valve, 2-Way (Straight Body)"),
    ("Body Material",    "Brass"),
    ("Port Size",        '1/4" NPT (Female both ends)'),
    ("Pressure Rating",  "400 psi"),
    ("Handle Type",      "Knurled Wheel Handle"),
    ("Media",            "Water (and compatible media)"),
    ("OEM",              "Acceptable per listing"),
]

for label, value in specs:
    ws.row_dimensions[row].height = 22
    cell(ws, row, 2, label, bold=True, size=10, bg=mid,  border_on=True,
         align="right")
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row,   end_column=4)
    cell(ws, row, 3, value, size=10, bg=light, border_on=True, wrap=True)
    row += 1

# ── PRODUCT IMAGE ─────────────────────────────────────────────────────────────
img_path = os.path.join(os.path.dirname(__file__), "prod01.png")
if os.path.exists(img_path):
    row += 1  # small gap
    ws.row_dimensions[row].height   = 130
    ws.row_dimensions[row+1].height = 8
    img = XLImage(img_path)
    img.width, img.height = 110, 110
    # anchor to column C of current row
    ws.add_image(img, f"C{row}")
    cell(ws, row, 2, "Product\nImage", bold=True, size=9, bg=mid,
         align="right", wrap=True, border_on=True)
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row,   end_column=4)
    ws.cell(row=row, column=3).fill   = fill(light)
    ws.cell(row=row, column=3).border = border()
    row += 2

# ── SECTION: INQUIRY ─────────────────────────────────────────────────────────
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
c = ws.cell(row=row, column=2, value="  INQUIRY")
c.font      = Font(bold=True, size=11, color=white, name="Calibri")
c.fill      = fill(gold)
c.alignment = Alignment(horizontal="left", vertical="center")
row += 1

questions = [
    ("Q1 – UL Listing",
     "Is this product currently UL Listed (Underwriters Laboratories)?\n"
     "If yes, please provide the UL File Number and the applicable "
     "UL Standard (e.g., UL 429, UL 1963, or equivalent)."),

    ("Q2 – Certificate",
     "Can you share a copy of the UL Certificate of Compliance or "
     "a test report from a NRTL (Nationally Recognized Testing "
     "Laboratory) for this valve?"),

    ("Q3 – Alternative\nCertifications",
     "If UL Listing is not available, does the product hold any "
     "equivalent North American certification such as CSA, ETL, or "
     "NSF? Please specify the standard and scope."),

    ("Q4 – MOQ & Lead Time",
     "What is the Minimum Order Quantity (MOQ) and standard lead "
     "time for this item?"),
]

for label, value in questions:
    ws.row_dimensions[row].height = 60
    cell(ws, row, 2, label, bold=True, size=9, bg=mid,  border_on=True,
         align="right", wrap=True)
    ws.merge_cells(start_row=row, start_column=3,
                   end_row=row,   end_column=4)
    cell(ws, row, 3, value, size=10, bg=light, border_on=True, wrap=True)
    row += 1

# ── FOOTER ───────────────────────────────────────────────────────────────────
row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
c = ws.cell(row=row, column=2,
            value="Please reply to this email with the requested documentation. "
                  "Thank you for your assistance.")
c.font      = Font(italic=True, size=9, color="555555", name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30

row += 1
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
ws.cell(row=row, column=2).fill = fill(navy)
ws.row_dimensions[row].height  = 6

# ── FREEZE & SAVE ─────────────────────────────────────────────────────────────
ws.freeze_panes = "B4"
ws.sheet_view.showGridLines = False

out_path = os.path.join(os.path.dirname(__file__),
                        "UL_Certification_Inquiry.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
